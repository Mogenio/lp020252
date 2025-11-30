# Calcular a média de cada aluno;
# Exibir os alunos aprovados e reprovados (média >= 6 para aprovação);
# Permitir salvar um novo arquivo Excel com uma coluna adicional chamada “Situação” (Aprovado/Reprovado).
# O usuário deve escolher o nome do arquivo de saída no console.
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\gabri\OneDrive\TRABALHOTURMA\MEDIATURMA.xlsx')
sheet = wb.active

sheet.cell(row=1, column=4, value="MÉDIA")
sheet.cell(row=1, column=5, value="SITUAÇÃO")

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=3):
    celula_nome, celula_n1, celula_n2 = row
    
    n1 = celula_n1.value
    n2 = celula_n2.value

    if isinstance(n1, (int, float)) and isinstance(n2, (int, float)):
        media = (n1 + n2) / 2
    else:
           media = 0
    
    situacao = "APROVADO" if media >= 6 else "REPROVADO"

    numero_linha = row[0].row
    sheet.cell(row=numero_linha, column=4, value=media)
    sheet.cell(row=numero_linha, column=5, value=situacao)


nome_arquivo = input("Digite o nome para salvar o arquivo (sem extensão): ")

caminho_completo = r'C:\Users\gabri\OneDrive\TRABALHOTURMA\\' + nome_arquivo + '.xlsx'

wb.save(caminho_completo)
print(f"Arquivo salvo com sucesso em: {caminho_completo}")
